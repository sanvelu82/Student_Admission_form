import tkinter
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from random import randint, choice
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl,xlrd
from openpyxl.styles import Border, Side,PatternFill,Font
from openpyxl import Workbook
import pathlib
import os
#------------------------------------------------------
def pdf():
    sheet1['F7'] = "Sandeep"
def Submit():
    global my_img
    global ogol
    global a1,n1,c1,s1,db1,aa,mt,re,co,bg,do,cn,dis,pin,f1,fo1,m1,mo1,ai1,mb1,sn,dj,g1
    a1=t12.get()
    n1=t1.get()
    c1=cls.get()
    s1=sec.get()
    db1=t3.get()
    aa=t2.get()
    mt=lan.get()
    re=rel.get()
    co=com.get()
    bg=bld.get()
    do=t7.get()
    cn=t9.get()
    dis=t10.get()
    pin=t11.get()
    try:
        g1=sex
    except:
        messagebox.showerror("error","Select Gender!")
    f1=t4.get()
    fo1=foc.get()
    m1=t5.get()
    mo1=moc.get()
    ai1=pai.get()
    mb1=t6.get()
    sn=t8.get()
    dj=t13.get()
    if a1=="Enter Admission Number" or n1=="Enter the Name" or c1==" Select Class" or s1==" Select Section" or db1=="dd-mm-yyyy" or f1=="Enter Father Name" or m1=="Enter Mother Name" or ai1==" Select Annual Income" or mb1=="Enter Mobile Number" or sn=="Enter Street Name" or dj=="dd-mm-yyyy" or aa=="Enter Aadhaar Number" or mt==" Choose Mother Tongue" or re==" Select Religion" or co==" Select Community" or bg==" Choose blood group" or do=="Enter Door No" or cn=="Enter City Name" or dis=="Enter District" or pin=="Enter Pincode" or mt=="" or re=="" or co=="" or bg=="" or ai1=="" or c1=="" or s1=="":
        messagebox.showerror("error","Few Deatails is Missing !")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=a1.upper())
        sheet.cell(column=2,row=sheet.max_row,value=n1.upper())
        sheet.cell(column=3,row=sheet.max_row,value=c1.upper())
        sheet.cell(column=4,row=sheet.max_row,value=s1.upper())
        sheet.cell(column=5,row=sheet.max_row,value=db1)
        sheet.cell(column=6,row=sheet.max_row,value=g1.upper())
        sheet.cell(column=7,row=sheet.max_row,value=f1.upper())
        sheet.cell(column=8,row=sheet.max_row,value=m1.upper())
        sheet.cell(column=9,row=sheet.max_row,value=re.upper())
        sheet.cell(column=10,row=sheet.max_row,value=co.upper())
        sheet.cell(column=11,row=sheet.max_row,value=mb1)
        sheet.cell(column=12,row=sheet.max_row,value=bg.capitalize())
        sheet.cell(column=13,row=sheet.max_row,value=dj)
        sheet.cell(column=14,row=sheet.max_row,value=do)
        sheet.cell(column=15,row=sheet.max_row,value=sn.upper())
        sheet.cell(column=16,row=sheet.max_row,value=cn.upper())
        sheet.cell(column=17,row=sheet.max_row,value=dis.upper())
        sheet.cell(column=18,row=sheet.max_row,value=pin)
        sheet.cell(column=19,row=sheet.max_row,value=aa)
        sheet.cell(column=20,row=sheet.max_row,value=ai1)
        sheet.cell(column=21,row=sheet.max_row,value=fo1.upper())
        sheet.cell(column=22,row=sheet.max_row,value=mo1.upper())
        sheet.cell(column=23,row=sheet.max_row,value=mt.upper())
        font1=Font(name='Cambria')
        for row in sheet.iter_rows():
            for cell in row:
                cell.font=font1
        file.save('Student_data.xlsx')
        messagebox.showinfo("Alert","Student Admitted to School, Click Ok to Get Admission Certificate")
        svv.withdraw()
        top=Toplevel()
        Label(top,background="white",image=logo3).place(x=10,y=10)
        top.geometry("995x700+260+45")
        top.config(bg="white")
        top.title("Student Admission  Certificate")
        Label(top, text="SRI VINAYAGA VIDYALAYA", font=("cambria", 15), fg='darkblue', bg='white').place(x=143, y=40)
        Label(top, text=" NURSERY & PRIMARY SCHOOL", font=("cambria", 12), fg='darkblue', bg='white').place(x=141,y=68)
        Label(top, text="  ADMISSION FORM", font=("cambria", 15), fg='blue', bg='white').place(x=383, y=120)
        color = "black"
        thickness = 1
        frame=Frame(top, highlightbackground="black", highlightthickness=thickness, height=445, width=930,
                      bg="white").place(x=30, y=170)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=215)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=255)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=300)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=345)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=390)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=435)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=480)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=525)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=1, width=930, bg="white").place(
            x=30, y=570)
        Frame(top, highlightbackground="black", highlightthickness=thickness, height=445, width=1, bg="white").place(
            x=480, y=170)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=445, width=1, bg="white").place(
            x=230, y=170)
        Frame(top, highlightbackground=color, highlightthickness=thickness, height=445, width=1, bg="white").place(
            x=700, y=170)
        lacolor = "green"
        Label(top, text="Name of the Student", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=180)
        Label(top, text="Gender", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=180)
        Label(top, text="Date of Birth", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=223)
        Label(top, text="Blood Group", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=223)
        Label(top, text="Father Name", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=266)
        Label(top, text="Father Occupation", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=266)
        Label(top, text="Mother Name", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=309)
        Label(top, text="Mother Occupation", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=309)
        Label(top, text="Class ", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=353)
        Label(top, text="Section ", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=353)
        Label(top, text="Admission Number ", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=398)
        Label(top, text="Date of Admission ", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=398)
        Label(top, text="Mobile No ", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=443)
        Label(top, text="Aadhaar Number ", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=443)
        Label(top, text="Religion ", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=491)
        Label(top, text="Community ", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=491)
        Label(top, text="Village Name", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=536)
        Label(top, text="City Name", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=536)
        Label(top, text="Mother Tongue ", font=("cambria", 13), fg=lacolor, bg="white").place(x=39, y=580)
        Label(top, text="Pin Code ", font=("cambria", 13), fg=lacolor, bg="white").place(x=490, y=580)
        ecolor = "black"
        # -------------------------------------------------------------------------------------
        name = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        name.place(x=240, y=183)
        name.insert(0, t1.get().upper())
        # ------------------------------------------------------------------------------------
        gen = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        gen.place(x=712, y=183)
        gen.insert(0, sex.upper())
        # -------------------------------------------------------------------------------------
        date1 = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        date1.place(x=240, y=224)
        date1.insert(0, t3.get())
        # ------------------------------------------------------------------------------------
        blodgr = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        blodgr.place(x=712, y=224)
        blodgr.insert(0, bld.get())
        # ------------------------------------------------------------------------------------
        fname = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        fname.place(x=240, y=267)
        fname.insert(0, t4.get().upper())
        # ------------------------------------------------------------------------------------
        focc = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        focc.place(x=712, y=267)
        focc.insert(0, foc.get().upper())
        # ------------------------------------------------------------------------------------
        mname = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        mname.place(x=240, y=312)
        mname.insert(0, t5.get().upper())
        # ------------------------------------------------------------------------------------
        mocc = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        mocc.place(x=712, y=312)
        mocc.insert(0, moc.get().upper())
        # ------------------------------------------------------------------------------------
        clas = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        clas.place(x=240, y=357)
        clas.insert(0, cls.get().upper())
        # ------------------------------------------------------------------------------------
        sect = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        sect.place(x=712, y=357)
        sect.insert(0, sec.get().upper())
        # ------------------------------------------------------------------------------------
        adno = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        adno.place(x=240, y=402)
        adno.insert(0, t12.get())
        # ------------------------------------------------------------------------------------
        dtad = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        dtad.place(x=712, y=402)
        dtad.insert(0, t13.get())
        # ------------------------------------------------------------------------------------
        mbno = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        mbno.place(x=240, y=447)
        mbno.insert(0, t6.get())
        # ------------------------------------------------------------------------------------
        aadhno = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        aadhno.place(x=712, y=447)
        aadhno.insert(0, t2.get())
        # ------------------------------------------------------------------------------------
        religion = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        religion.place(x=240, y=492)
        religion.insert(0, rel.get().upper())
        # ------------------------------------------------------------------------------------
        community = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        community.place(x=712, y=492)
        community.insert(0, com.get().upper())
        # ------------------------------------------------------------------------------------
        village = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        village.place(x=240, y=537)
        village.insert(0, t8.get().upper())
        # ------------------------------------------------------------------------------------
        city = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        city.place(x=712, y=537)
        city.insert(0, t9.get().upper())
        # ------------------------------------------------------------------------------------
        motherton = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        motherton.place(x=240, y=582)
        motherton.insert(0, lan.get().upper())
        # ------------------------------------------------------------------------------------
        pincode = Entry(top, border=0, width=26, font=("cambria", 13), fg=ecolor)
        pincode.place(x=712, y=582)
        pincode.insert(0, t11.get().upper())
        # ------------------------------------------------------------------------------------
        print_pdf = Button(top, bg="#275BD2", border=0, text="  Print PDF  ", font=("cambria", 13), fg="white",command=pdf).place(x=800, y=40)
        top.overrideredirect(1)
        top.mainloop()
#------------------------------------------------------
sex=""
def selection():
    global sex
    value=gender.get()
    if value==1:
        sex="Male"
        print("Male")
    else:
        sex="Female"
        print("Female")
svv=Tk()
svv.title("Student Admission Form")
svv.geometry("1250x700+130+45")
svv.config(background="white")
log=PhotoImage(file="C:\\Users\\SANDEEP\\Documents\\Student Logo.png")
logo=Image.open("C:\\Users\\SANDEEP\\Documents\\svv-removebg-preview.png")
logo3=logo.resize((120,95))
logo=logo.resize((150,115))
logo3=ImageTk.PhotoImage(logo3)
logo=ImageTk.PhotoImage(logo)
svv.iconphoto(False,log)
Label(svv,image=logo,bg="white").place(x=10,y=10)
#----------------------------------------------------------------------------------------------
file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    border_style=Border(left=Side(style="thin"),
                  right=Side(style="thin"),
                  top=Side(style="thin"),
                  bottom=Side(style="thin"))
    cell_color=PatternFill(patternType='solid',fgColor='FFFF00')
    sheet['A1']="Admission No"
    sheet.column_dimensions['A'].width=13
    sheet['A1'].fill=cell_color
    sheet['B1'] ="Name of the Student"
    sheet.column_dimensions['B'].width = 25
    sheet['B1'].fill = cell_color
    sheet['C1'] ="Class"
    sheet.column_dimensions['C'].width = 8
    sheet['C1'].fill = cell_color
    sheet['D1'] ="Section"
    sheet.column_dimensions['D'].width = 10
    sheet['D1'].fill = cell_color
    sheet['E1'] ="Date of Birth"
    sheet.column_dimensions['E'].width = 13
    sheet['E1'].fill = cell_color
    sheet['F1'] ="Gender"
    sheet.column_dimensions['F'].width = 8
    sheet['F1'].fill = cell_color
    sheet['G1'] ="Father Name"
    sheet.column_dimensions['G'].width = 19
    sheet['G1'].fill = cell_color
    sheet['H1'] ="Mother Name"
    sheet.column_dimensions['H'].width = 19
    sheet['H1'].fill = cell_color
    sheet['I1'] ="Religion"
    sheet.column_dimensions['I'].width = 9
    sheet['I1'].fill = cell_color
    sheet['J1'] ="Community"
    sheet.column_dimensions['J'].width = 12
    sheet['J1'].fill = cell_color
    sheet['K1'] ="Mobile Number"
    sheet.column_dimensions['K'].width = 16
    sheet['K1'].fill = cell_color
    sheet['L1'] ="Blood Group"
    sheet.column_dimensions['L'].width = 12
    sheet['L1'].fill = cell_color
    sheet['M1'] ="Date of joining"
    sheet.column_dimensions['M'].width = 14
    sheet['M1'].fill = cell_color
    sheet['N1'] ="Door No"
    sheet.column_dimensions['N'].width = 8.5
    sheet['N1'].fill = cell_color
    sheet['O1'] ="Street Name"
    sheet.column_dimensions['O'].width = 15
    sheet['O1'].fill = cell_color
    sheet['P1'] ="City Name"
    sheet.column_dimensions['P'].width = 15
    sheet['P1'].fill = cell_color
    sheet['Q1'] ="District"
    sheet.column_dimensions['Q'].width = 15
    sheet['Q1'].fill = cell_color
    sheet['R1'] ="Pincode"
    sheet.column_dimensions['R'].width = 10
    sheet['R1'].fill = cell_color
    sheet['S1'] ="Aadhaar Number"
    sheet.column_dimensions['S'].width = 17
    sheet['S1'].fill = cell_color
    sheet['T1'] ="Parent's Income"
    sheet.column_dimensions['T'].width = 15
    sheet['T1'].fill = cell_color
    sheet['U1'] ="Father Occupation"
    sheet.column_dimensions['U'].width = 18
    sheet['U1'].fill = cell_color
    sheet['V1'] ="Mother Occupation"
    sheet.column_dimensions['V'].width = 18
    sheet['V1'].fill = cell_color
    sheet['W1'] ="Mother Tongue"
    sheet.column_dimensions['W'].width = 15
    sheet['W1'].fill = cell_color
    sheet['A1'].border = border_style
    sheet['B1'].border = border_style
    sheet['C1'].border = border_style
    sheet['D1'].border = border_style
    sheet['E1'].border = border_style
    sheet['F1'].border = border_style
    sheet['G1'].border = border_style
    sheet['H1'].border = border_style
    sheet['I1'].border = border_style
    sheet['J1'].border = border_style
    sheet['K1'].border = border_style
    sheet['L1'].border = border_style
    sheet['M1'].border = border_style
    sheet['N1'].border = border_style
    sheet['I1'].border = border_style
    sheet['J1'].border = border_style
    sheet['K1'].border = border_style
    sheet['L1'].border = border_style
    sheet['M1'].border = border_style
    sheet['N1'].border = border_style
    sheet['O1'].border = border_style
    sheet['P1'].border = border_style
    sheet['Q1'].border = border_style
    sheet['R1'].border = border_style
    sheet['S1'].border = border_style
    sheet['T1'].border = border_style
    sheet['U1'].border = border_style
    sheet['V1'].border = border_style
    sheet['W1'].border = border_style
    font1 = Font(name='Cambria')
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font1
    global sheet1
    sheet1 = file.create_sheet('Sheet1')
    sheet1.merge_cells("M9:Q9")
    file.save('Student_data.xlsx')

#-----------------------------------------------------------------------------------
logo1=Image.open("C:\\Users\\SANDEEP\\Documents\\child.jpg")
logo1=logo1.resize((125,115))
logo1=ImageTk.PhotoImage(logo1)
Label(svv,image=logo1,bg="white").place(x=1100,y=10)
Label(svv,text="' Let us remember : One book ,",font=("cambria",15),fg='darkblue',bg='white').place(x=1233,y=30)
Label(svv,text="   One Pen, One Child, One",font=("cambria",15),fg='darkblue',bg='white').place(x=1233,y=58)
Label(svv,text="teacher can change the world ! '",font=("cambria",15),fg='darkblue',bg='white').place(x=1233,y=87)
#-----------------------------------------------------------------------------------
Label(svv,text="SRI VINAYAGA VIDYALAYA",font=("cambria",18),fg='darkblue',bg='white').place(x=163,y=40)
Label(svv,text=" NURSERY & PRIMARY SCHOOL",font=("cambria",14),fg='darkblue',bg='white').place(x=163,y=75)
Label(svv,text="STUDENTS ADMISSION FORM",font=("cambria",17),fg='darkblue',bg='white').place(x=565,y=105)
Label(svv,text="Name of the Student",fg="#474646",bg="white",font=("cambria",15)).place(x=20,y=180)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=203,y=180)
Label(svv,text="AADHAAR Number",fg="#474646",bg="white",font=("cambria",15)).place(x=400,y=180)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=569,y=180)
Label(svv,text="Date of Birth",fg="#474646",bg="white",font=("cambria",15)).place(x=800,y=180)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=915,y=180)
Label(svv,text="Mother Tongue",fg="#474646",bg="white",font=("cambria",15)).place(x=1200,y=180)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=1340,y=180)
Label(svv,text="Gender",fg="#474646",bg="white",font=("cambria",15)).place(x=20,y=270)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=90,y=270)
Label(svv,text="Religion",fg="#474646",bg="white",font=("cambria",15)).place(x=400,y=270)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=480,y=270)
Label(svv,text="Community",fg="#474646",bg="white",font=("cambria",15)).place(x=800,y=270)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=910,y=270)
Label(svv,text="Blood Group",fg="#474646",bg="white",font=("cambria",15)).place(x=1200,y=270)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=1310,y=270)
Label(svv,text="Father Name",fg="#474646",bg="white",font=("cambria",15)).place(x=20,y=360)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=140,y=360)
Label(svv,text="Father's Occupation",fg="#474646",bg="white",font=("cambria",15)).place(x=400,y=360)
Label(svv,text="Mother Name",fg="#474646",bg="white",font=("cambria",15)).place(x=800,y=360)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=925,y=360)
Label(svv,text="Mother's Occupation",fg="#474646",bg="white",font=("cambria",15)).place(x=1200,y=360)
Label(svv,text="Parent's Annual Income",fg="#474646",bg="white",font=("cambria",15)).place(x=20,y=450)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=230,y=450)
Label(svv,text="Mobile Number",fg="#474646",bg="white",font=("cambria",15)).place(x=400,y=450)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=540,y=450)
Label(svv,text="Door No",fg="#474646",bg="white",font=("cambria",15)).place(x=800,y=450)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=880,y=450)
Label(svv,text="Street Name",fg="#474646",bg="white",font=("cambria",15)).place(x=1200,y=450)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=1310,y=450)
Label(svv,text="City Name",fg="#474646",bg="white",font=("cambria",15)).place(x=20,y=540)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=115,y=540)
Label(svv,text="District",fg="#474646",bg="white",font=("cambria",15)).place(x=400,y=540)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=470,y=540)
Label(svv,text="Pincode",fg="#474646",bg="white",font=("cambria",15)).place(x=800,y=540)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=874,y=540)
Label(svv,text="Date of Joining",fg="#474646",bg="white",font=("cambria",15)).place(x=1200,y=540)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=1330,y=540)
Label(svv,text="Class for Which Admission is sought for",fg="#474646",bg="white",font=("cambria",15)).place(x=20,y=630)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=365,y=630)
Label(svv,text="Section",fg="#474646",bg="white",font=("cambria",15)).place(x=400,y=630)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=470,y=630)
Label(svv,text="Admission Number",fg="#474646",bg="white",font=("cambria",15)).place(x=800,y=630)
Label(svv,text="*",fg="Red",bg="white",font=("cambria",12)).place(x=974,y=630)
#------------------------------------------------------------------------------
def on_enter(e):
    n = t1.get()
    if n == 'Enter the Name':
        t1.delete(0,'end')
        t1.config(fg="black")
def on_leave(e):
    name=t1.get()
    if name=='':
        t1.insert(0,'Enter the Name')
        t1.config(fg='#777777')
t1=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t1.place(x=25,y=220)
t1.insert(0,"Enter the Name")
t1.bind('<FocusIn>', on_enter)
t1.bind('<FocusOut>', on_leave)
#------------------------------------------------------------------------------------------
def on_enter(e):
    n=t2.get()
    if n=='Enter Aadhaar Number':
        t2.delete(0,'end')
        t2.config(fg="black")
def on_leave(e):
    name=t2.get()
    if name=='':
        t2.insert(0,'Enter Aadhaar Number')
        t2.config(fg='#777777')

t2=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t2.place(x=405,y=220)
t2.insert(0,"Enter Aadhaar Number")
t2.bind('<FocusIn>', on_enter)
t2.bind('<FocusOut>', on_leave)
#----------------------------------------------------------------------------
def on_enter(e):
    n=t3.get()
    if n=='dd-mm-yyyy':
        t3.delete(0,'end')
        t3.config(fg="black")
def on_leave(e):
    name=t3.get()
    if name=='':
        t3.insert(0,'dd-mm-yyyy')
        t3.config(fg='#777777')
t3=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t3.place(x=807,y=220)
t3.insert(0,"dd-mm-yyyy")
t3.bind('<FocusIn>', on_enter)
t3.bind('<FocusOut>', on_leave)
#----------------------------------------------------------------------------
lan=ttk.Combobox(svv,width=20,font=("cambria",15))
lan.place(x=1205,y=218)
lan['values']=(" Tamil"," Telugu"," Kannada"," Malayalam")
lan.insert(0," Choose Mother Tongue")
#----------------------------------------------------------------------------
gender=IntVar()
male=Radiobutton(svv,text = "Male",bg='white',variable=gender,value=1,font='cambria',command=selection)
male.place(x=30,y=300)
female=Radiobutton(svv,text = 'Female',bg='white',variable=gender,value=2,font='cambria',command=selection)
female.place(x=160,y=300)
#----------------------------------------------------------------------------
rel=ttk.Combobox(svv,width=20,font=("cambria",15))
rel.place(x=405,y=308)
rel['values']=(" Hindu"," Christian"," Muslim"," Jainism", " Buddhism", " Others")
rel.insert(0," Select Religion")
#----------------------------------------------------------------------------
com=ttk.Combobox(svv,width=20,font=("cambria",15))
com.place(x=805,y=308)
com['values']=(" BC-Others"," MBC"," ST"," SC-Others", " OC", " DNC(Denotified Communities)", " No Community")
com.insert(0," Select Community")
#----------------------------------------------------------------------------
bld=ttk.Combobox(svv,width=20,font=("cambria",15))
bld.place(x=1205,y=308)
bld['values']=(" Don't Know"," A+ve"," A-ve"," B+ve"," B-ve", " AB+ve", " AB-ve", " O+ve", " O-ve", " A1+ve", " A1-ve")
bld.insert(0," Choose blood group")
#------------------------------------------------------------------------------
def on_enter(e):
    n = t4.get()
    if n == 'Enter Father Name':
        t4.delete(0,'end')
        t4.config(fg="black")
def on_leave(e):
    name=t4.get()
    if name=='':
        t4.insert(0,'Enter Father Name')
        t4.config(fg='#777777')
t4=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t4.place(x=25,y=400)
t4.insert(0,"Enter Father Name")
t4.bind('<FocusIn>', on_enter)
t4.bind('<FocusOut>', on_leave)
#----------------------------------------------------------------------------
foc=ttk.Combobox(svv,width=20,font=("cambria",15))
foc.place(x=405,y=400)
foc['values']=(" Government"," Private"," Self-employed"," Daily wages"," Un-employed", " N/A")
foc.insert(0," Select Occupation")
#----------------------------------------------------------------------------
def on_enter(e):
    n = t5.get()
    if n == 'Enter Mother Name':
        t5.delete(0,'end')
        t5.config(fg="black")
def on_leave(e):
    name=t5.get()
    if name=='':
        t5.insert(0,'Enter Mother Name')
        t5.config(fg='#777777')
t5=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t5.place(x=805,y=400)
t5.insert(0,"Enter Mother Name")
t5.bind('<FocusIn>', on_enter)
t5.bind('<FocusOut>', on_leave)
#----------------------------------------------------------------------------
moc=ttk.Combobox(svv,width=20,font=("cambria",15))
moc.place(x=1205,y=400)
moc['values']=(" Government"," Private"," Self-employed"," Daily wages"," Un-employed", " N/A")
moc.insert(0," Select Occupation")
#-----------------------------------------------------------------------------
pai=ttk.Combobox(svv,width=20,font=("cambria",15))
pai.place(x=25,y=487)
pai['values']=(" 0 to 12000"," 12001-24000"," 24001-50000"," 50001-100000"," Above 100000")
pai.insert(0," Select Annual Income")
#----------------------------------------------------------------------
def on_enter(e):
    n = t6.get()
    if n == 'Enter Mobile Number':
        t6.delete(0,'end')
        t6.config(fg="black")
def on_leave(e):
    name=t6.get()
    if name=='':
        t6.insert(0,'Enter Mobile Number')
        t6.config(fg='#777777')
t6=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t6.place(x=405,y=487)
t6.insert(0,"Enter Mobile Number")
t6.bind('<FocusIn>', on_enter)
t6.bind('<FocusOut>', on_leave)
#-----------------------------------------------------------------------------
def on_enter(e):
    n = t7.get()
    if n == 'Enter Door No':
        t7.delete(0,'end')
        t7.config(fg="black")
def on_leave(e):
    name=t7.get()
    if name=='':
        t7.insert(0,'Enter Door No')
        t7.config(fg='#777777')
t7=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t7.place(x=805,y=487)
t7.insert(0,"Enter Door No")
t7.bind('<FocusIn>', on_enter)
t7.bind('<FocusOut>', on_leave)
#-----------------------------------------------------------------------------
def on_enter(e):
    n = t8.get()
    if n == 'Enter Street Name':
        t8.delete(0,'end')
        t8.config(fg="black")
def on_leave(e):
    name=t8.get()
    if name=='':
        t8.insert(0,'Enter Street Name')
        t8.config(fg='#777777')
t8=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t8.place(x=1205,y=487)
t8.insert(0,"Enter Street Name")
t8.bind('<FocusIn>', on_enter)
t8.bind('<FocusOut>', on_leave)
#-----------------------------------------------------------------------------
def on_enter(e):
    n = t9.get()
    if n == 'Enter City Name':
        t9.delete(0,'end')
        t9.config(fg="black")
def on_leave(e):
    name=t9.get()
    if name=='':
        t9.insert(0,'Enter City Name')
        t9.config(fg='#777777')
t9=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t9.place(x=25,y=580)
t9.insert(0,"Enter City Name")
t9.bind('<FocusIn>', on_enter)
t9.bind('<FocusOut>', on_leave)
#------------------------------------------------------------
def on_enter(e):
    n = t10.get()
    if n == 'Enter District':
        t10.delete(0,'end')
        t10.config(fg="black")
def on_leave(e):
    name=t10.get()
    if name=='':
        t10.insert(0,'Enter District')
        t10.config(fg='#777777')
t10=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t10.place(x=405,y=580)
t10.insert(0,"Enter District")
t10.bind('<FocusIn>', on_enter)
t10.bind('<FocusOut>', on_leave)
#------------------------------------------------------------
def on_enter(e):
    n = t11.get()
    if n == 'Enter Pincode':
        t11.delete(0,'end')
        t11.config(fg="black")
def on_leave(e):
    name=t11.get()
    if name=='':
        t11.insert(0,'Enter Pincode')
        t11.config(fg='#777777')
t11=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t11.place(x=805,y=580)
t11.insert(0,"Enter Pincode")
t11.bind('<FocusIn>', on_enter)
t11.bind('<FocusOut>', on_leave)
#-------------------------------------------------
def on_enter(e):
    n=t13.get()
    if n=='dd-mm-yyyy':
        t13.delete(0,'end')
        t13.config(fg="black")
def on_leave(e):
    name=t13.get()
    if name=='':
        t13.insert(0,'dd-mm-yyyy')
        t13.config(fg='#777777')
t13=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t13.place(x=1205,y=580)
t13.insert(0,"dd-mm-yyyy")
t13.bind('<FocusIn>', on_enter)
t13.bind('<FocusOut>', on_leave)
#-----------------------------------------------------------------------------
cls=ttk.Combobox(svv,width=20,font=("cambria",15))
cls.place(x=25,y=670)
cls['values']=(" LKG"," UKG"," I"," II"," III"," IV"," V")
cls.insert(0," Select Class")
#---------------------------------------------------------------
sec=ttk.Combobox(svv,width=20,font=("cambria",15))
sec.place(x=405,y=670)
sec['values']=(" A"," B"," C"," D"," E")
sec.insert(0," Select Section")
#---------------------------------------------------------------
def on_enter(e):
    n = t12.get()
    if n == 'Enter Admission Number':
        t12.delete(0,'end')
        t12.config(fg="black")
def on_leave(e):
    name=t12.get()
    if name=='':
        t12.insert(0,'Enter Admission Number')
        t12.config(fg='#777777')
t12=Entry(svv,fg="#777777",bg="white",border=1,width=24,font=("cambria",15))
t12.place(x=805,y=670)
t12.insert(0,"Enter Admission Number")
t12.bind('<FocusIn>', on_enter)
t12.bind('<FocusOut>', on_leave)
#---------------------------------------------------------------------
sub=Button(svv,text="   Submit   ",bg='green',fg='white',border=0,font=("cambria",15,"bold"),command=Submit).place(x=1200,y=670)
svv.state('zoomed')
svv.mainloop()
